from django.shortcuts import get_object_or_404

from .serializers import *
from django.db.models import Sum, Q
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from .models import *

from users.models import User


# Create your views here.

class UserCountAPIView(APIView):
    def get(self, request, format=None):
        total_users = User.objects.count()
        if request.user.is_authenticated:
            try:
                statistic = Statistic.objects.get(user=request.user)
            except Statistic.DoesNotExist:
                statistic = Statistic(user=request.user, count=1)
                statistic.save()
            else:
                statistic.count += 1
                statistic.save()
        registered_users_count = Statistic.objects.aggregate(total_registered_users=Sum('count'))[
            'total_registered_users']

        return Response({
            'total_users': total_users
        })


class LegalDocuments(APIView):
    def get(self, request, format=None):
        legal_documents = Legal_Documents.objects.all()
        serializer = Legal_DocumentsSerializer(legal_documents, many=True)
        return Response(serializer.data)


class NewsList(APIView):
    def get(self, request, format=None):
        news = News.objects.all()
        serializer = NewsSerializer(news, many=True)
        return Response(serializer.data)


class NewsDetail(APIView):
    def get(self, request, pk, format=None):
        news = News.objects.get(pk=pk)
        serializer = NewsSerializer(news)
        return Response(serializer.data)


class BannerList(APIView):
    def get(self, request, format=None):
        banners = Banner.objects.all()
        serializer = BannerSerializer(banners, many=True)
        return Response(serializer.data)



def create_sheet_with_data(wb, title, df):
    ws = wb.create_sheet(title=title)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            if r_idx == 1:
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

class CalculateCostAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            regions = Area.objects.all()
            region_list = [{'id': region.id, 'name': region.name} for region in regions]

            seeds = Seed.objects.all()
            seed_list = [{'id': seed.id, 'name': seed.name, 'area': seed.area.name} for seed in seeds]

            trees = Tree.objects.all()
            tree_list = [{'id': tree.id, 'name': tree.name, 'area': tree.area.name} for tree in trees]

            response_data = {
                'regions': region_list,
                'plant_types': ['seed', 'tree'],
                'seeds': seed_list,
                'trees': tree_list
            }

            return Response(response_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        try:
            region = request.data.get('region')
            plant_type = request.data.get('plant_type')
            plant_name = request.data.get('plant_name')
            areas_to_sow = request.data.get('areas_to_sow')
            unit = request.data.get('unit', 'hectare')
            user_id = request.data.get('user_id')

            if not region or not plant_type or not plant_name or not areas_to_sow or not user_id:
                return Response({'error': 'Регион, тип растения, имя растения, площадь для посева и ID пользователя обязательны.'},
                                status=status.HTTP_400_BAD_REQUEST)

            areas_to_sow = float(areas_to_sow)
            if unit == 'sotix':
                areas_to_sow /= 100  # 1 gektar = 100 sotix

            area = Area.objects.filter(name=region).first()
            if not area:
                return Response({'error': 'Регион не существует'}, status=status.HTTP_400_BAD_REQUEST)

            user = User.objects.filter(id=user_id).first()
            if not user:
                return Response({'error': 'Пользователь не существует'}, status=status.HTTP_400_BAD_REQUEST)

            total_cost = 0

            categories = []
            unit_prices = []
            costs_per_unit = []
            total_costs = []

            if plant_type == 'seed':
                seed = Seed.objects.filter(name=plant_name, area=area).first()
                if not seed:
                    return Response({'error': 'Семя не существует в указанном регионе'}, status=status.HTTP_400_BAD_REQUEST)

                seed_cost = seed.price * areas_to_sow
                total_cost += seed_cost

                categories.append('Семена')
                unit_prices.append(seed.price)
                costs_per_unit.append(seed.price)
                total_costs.append(seed_cost)

                workers = Worker.objects.filter(area=area, seed=seed)
                tractors = Tractor.objects.filter(area=area, seed=seed)

            elif plant_type == 'tree':
                tree = Tree.objects.filter(name=plant_name, area=area).first()
                if not tree:
                    return Response({'error': 'Дерево не существует в указанном регионе'}, status=status.HTTP_400_BAD_REQUEST)

                spacing = 4
                trees_per_hectare = (10000 / (spacing * spacing))
                trees_needed = trees_per_hectare * areas_to_sow

                tree_cost = tree.price * trees_needed
                total_cost += tree_cost

                categories.append('Деревья')
                unit_prices.append(tree.price)
                costs_per_unit.append(tree.price * trees_per_hectare)
                total_costs.append(tree_cost)

                workers = Worker.objects.filter(area=area, tree=tree)
                tractors = Tractor.objects.filter(area=area, tree=tree)

            else:
                return Response({'error': 'Неправильный тип растения'}, status=status.HTTP_400_BAD_REQUEST)

            for worker in workers:
                worker_assignments = WorkerTariffAssignment.objects.filter(worker=worker)
                for assignment in worker_assignments:
                    tariff = assignment.tariff
                    hours = assignment.hours
                    worker_total_cost = tariff.wage_for_7_hours * hours * areas_to_sow
                    categories.append(f'Рабочие (Разряд {tariff.name})')
                    unit_prices.append(tariff.wage_for_7_hours)
                    costs_per_unit.append(tariff.wage_for_7_hours * hours)
                    total_costs.append(worker_total_cost)
                    total_cost += worker_total_cost

            for tractor in tractors:
                tractor_assignments = TractorTariffAssignment.objects.filter(tractor=tractor)
                for assignment in tractor_assignments:
                    tariff = assignment.tariff
                    hours = assignment.hours
                    tractor_total_cost = tariff.wage_for_7_hours * hours * areas_to_sow
                    categories.append(f'Тракторы (Разряд {tariff.name})')
                    unit_prices.append(tariff.wage_for_7_hours)
                    costs_per_unit.append(tariff.wage_for_7_hours * hours)
                    total_costs.append(tractor_total_cost)
                    total_cost += tractor_total_cost
            fertilisers = Fertiliser.objects.filter(area=area, seed=seed) if plant_type == 'seed' else Fertiliser.objects.filter(area=area, tree=tree)
            fertiliser_needed_per_hectare = 100
            fertiliser_total_cost = 0

            for fertiliser in fertilisers:
                fertiliser_cost_per_kg = fertiliser.price_per_kg
                fertiliser_cost_per_unit = fertiliser_cost_per_kg * fertiliser_needed_per_hectare
                total_fertiliser_cost = fertiliser_cost_per_unit * areas_to_sow
                fertiliser_total_cost += total_fertiliser_cost

                categories.append(f'Удобрения ({fertiliser.name})')
                unit_prices.append(fertiliser.price_per_kg)
                costs_per_unit.append(fertiliser_cost_per_unit)
                total_costs.append(total_fertiliser_cost)

            total_cost += fertiliser_total_cost

            # Work types
            work_types = WorkType.objects.filter(Q(worker_tariff__isnull=False) | Q(tractor_tariff__isnull=False))
            work_type_total_cost = 0

            for work_type in work_types:
                work_type_hours = work_type.hours
                if work_type.worker_tariff:
                    worker_cost = work_type.worker_tariff.wage_for_7_hours * work_type_hours * areas_to_sow
                    categories.append(f'Тип работы (Рабочий, Разряд {work_type.worker_tariff.name})')
                    unit_prices.append(work_type.worker_tariff.wage_for_7_hours)
                    costs_per_unit.append(work_type.worker_tariff.wage_for_7_hours * work_type_hours)
                    total_costs.append(worker_cost)
                    total_cost += worker_cost

                if work_type.tractor_tariff:
                    tractor_cost = work_type.tractor_tariff.wage_for_7_hours * work_type_hours * areas_to_sow
                    categories.append(f'Тип работы (Трактор, Разряд {work_type.tractor_tariff.name})')
                    unit_prices.append(work_type.tractor_tariff.wage_for_7_hours)
                    costs_per_unit.append(work_type.tractor_tariff.wage_for_7_hours * work_type_hours)
                    total_costs.append(tractor_cost)
                    total_cost += tractor_cost

            categories.append('Итого')
            unit_prices.append('')
            costs_per_unit.append('')
            total_costs.append(total_cost)

            summary_data = {
                'Категория': categories,
                'За единицу': unit_prices,
                'Стоимость за единицу': costs_per_unit,
                'Общая стоимость': total_costs
            }
            summary_df = pd.DataFrame(summary_data)

            excel_file = io.BytesIO()
            wb = Workbook()
            summary_ws = wb.active
            summary_ws.title = "Сводка"

            for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))
                    if r_idx == 1:
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            for column in summary_ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                summary_ws.column_dimensions[column[0].column_letter].width = adjusted_width

            create_sheet_with_data(wb, "Сводка", summary_df)

            wb.save(excel_file)
            excel_file.seek(0)
            file_name = 'cost_estimation.xlsx'
            path = default_storage.save(f'excel_files/{file_name}', ContentFile(excel_file.getvalue()))

            order = Order.objects.create(
                user=user,
                excel_file=path
            )

            return Response({'file_path': path, 'order_id': order.id}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MemsList(APIView):
    def get(self, request, format=None):
        mems = Mems.objects.all()
        serializer = MemsSerializer(mems, many=True)
        return Response(serializer.data)


class MemsDetails(APIView):

    def get(self, request, pk, format=None):
        mem = get_object_or_404(Mems, pk=pk)
        serializer = MemsSerializer(mem)
        return Response(serializer.data)

    def put(self, request, pk, format=None):
        mem = get_object_or_404(Mems, pk=pk)
        serializer = MemsSerializer(mem, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, format=None):
        mem = get_object_or_404(Mems, pk=pk)
        mem.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)